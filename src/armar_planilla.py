# -*- coding: utf-8 -*-
"""
Arma una planilla de control de retenciones (.xlsx) a partir de un CSV de datos
por recibo. Genera 3 hojas: Resumen (teorico vs real), Datos y Parametros.

Uso:
    python armar_planilla.py  datos.csv  salida.xlsx  [opciones]

Opciones:
    --mes-siguiente / --mes-en-curso   modalidad de pago (default: mes-en-curso, el indice del CSV = mes de pago)
    --con-prorrateo-sac                prorratea 1/12 de la remuneracion como SAC (default: NO)
    --aportes-tope                     aportes = 17% s/ tope siempre (default: 17% s/ tope, real si es menor)
    --conyuge / --hijos / --hijo-incap habilita esas cargas de familia

Formato del CSV (delimitador ';', encabezado obligatorio, una fila por recibo):
    remun;sac;no_computable;aporte_real;sindical;otros_desc;alq_beneficio;
    ded_grales_otras;cajas_compl;conyuge;hijos;hijo_incap;ret_informada
Numeros en formato es-AR ("1.234.567,89") o simple ("1234567.89"). Celdas vacias = 0
(en ret_informada, vacio significa "usar la retencion teorica").

NO incluye datos de contribuyentes: el CSV lo arma el usuario y no se versiona.
"""
import csv
import sys

from engine import liquidar, dp_acumulada, MESES, TOPE_SS, SCALES, ABREV

CAMPOS = ["remun", "sac", "no_computable", "aporte_real", "sindical", "otros_desc",
          "alq_beneficio", "ded_grales_otras", "cajas_compl", "conyuge", "hijos",
          "hijo_incap", "ret_informada"]


def leer_csv(path):
    filas = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            filas.append(row)
    n = len(filas)

    def parse(s):
        s = (s or "").strip()
        if s == "":
            return None
        if "," in s:                       # formato es-AR: 1.234.567,89
            s = s.replace(".", "").replace(",", ".")
        return float(s)

    def col(nombre, permitir_none=False):
        out = []
        for r in filas:
            v = parse(r.get(nombre))
            out.append(v if v is not None else (None if permitir_none else 0.0))
        return out + [None] * (12 - n)

    datos = dict(
        remun=col("remun"), sac=col("sac"), no_computable=col("no_computable"),
        ap_jub=col("aporte_real"), ap_os=[0] * 12, ap_pami=[0] * 12,
        sindical=col("sindical"), otros_desc=col("otros_desc"),
        alq_beneficio=col("alq_beneficio"), ded_grales_otras=col("ded_grales_otras"),
        cajas_compl=col("cajas_compl"),
        conyuge=col("conyuge"), hijos=col("hijos"), hijo_incap=col("hijo_incap"),
        ret_informada=col("ret_informada", permitir_none=True),
    )
    return datos, n


def armar(csv_path, xlsx_path, **kw):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    datos, n = leer_csv(csv_path)
    flags = dict(aplica_conyuge=kw.get("conyuge", False),
                 aplica_hijos=kw.get("hijos", False),
                 aplica_hijo_incap=kw.get("hijo_incap", False))
    datos.update(flags)
    liq_kw = dict(modalidad_mes_siguiente=kw.get("mes_siguiente", False),
                  usar_aportes_reales=not kw.get("aportes_tope", False),
                  sac_prorrateo=kw.get("prorrateo_sac", False))

    teo = liquidar({**datos, "ret_informada": [None] * 12}, **liq_kw)   # teorico puro
    real = liquidar(datos, **liq_kw)                                    # con retenciones informadas

    H1 = Font(bold=True, size=14, color="1F3864")
    H2 = Font(bold=True, size=10, color="FFFFFF")
    BOLD = Font(bold=True)
    FILL = PatternFill("solid", fgColor="2E75B6")
    money = "#,##0.00"
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    R = Alignment(horizontal="right")
    WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")

    wb = openpyxl.Workbook()

    # ---- Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "CONTROL DE RETENCIONES - GANANCIAS 4a CATEGORIA"
    ws["B2"].font = H1
    ws["B3"] = f"Recibos: {n}   |   modalidad: {'MES SIGUIENTE' if liq_kw['modalidad_mes_siguiente'] else 'MES EN CURSO'}"
    cols = ["Recibo", "Mes pago", "Rem. bruta acum.", "Aportes acum.", "Ded. grales acum.",
            "Ded. personales acum.", "Gcia neta suj. imp.", "Imp. determ. acum.",
            "Ret. TEORICA", "Ret. REAL", "Dif."]
    for j, c in enumerate(cols):
        cc = ws.cell(5, 2 + j, c)
        cc.font = H2
        cc.fill = FILL
        cc.alignment = WRAP
    ws.row_dimensions[5].height = 40
    for i in range(n):
        t = teo[i]
        rv = datos["ret_informada"][i]
        rr = real[i]["ret_real"] if rv is not None else None
        vals = [i + 1, t["mes_pago"].capitalize(), t["rem_acum"], t["ap_acum"], t["ded_grales_acum"],
                t["dp_total"], t["gnsi"], t["imp_det"], t["ret_dev"],
                (rr if rr is not None else "s/dato"),
                (t["ret_dev"] - rr if rr is not None else None)]
        for j, v in enumerate(vals):
            cc = ws.cell(6 + i, 2 + j, v)
            cc.border = thin
            if isinstance(v, (int, float)) and j > 0:
                cc.number_format = money
                cc.alignment = R
    rr = 6 + n
    ws.cell(rr, 2, "TOTAL").font = BOLD
    c = ws.cell(rr, 10, sum(x["ret_dev"] for x in teo))
    c.number_format = money
    c.font = BOLD
    ndat = sum(1 for v in datos["ret_informada"][:n] if v is not None)
    if ndat:
        c = ws.cell(rr, 11, sum(v for v in datos["ret_informada"][:n] if v is not None))
        c.number_format = money
        c.font = BOLD
    for j, w in enumerate([9, 12, 17, 15, 16, 18, 18, 18, 15, 15, 14]):
        ws.column_dimensions[get_column_letter(2 + j)].width = w

    # ---- Datos ----
    wd = wb.create_sheet("Datos")
    wd.sheet_view.showGridLines = False
    wd["B2"] = "DATOS DE ENTRADA (del CSV)"
    wd["B2"].font = H1
    wd.cell(4, 2, "Recibo").font = H2
    wd.cell(4, 2).fill = FILL
    for j, campo in enumerate(CAMPOS):
        cc = wd.cell(4, 3 + j, campo)
        cc.font = H2
        cc.fill = FILL
    mapa = {"remun": "remun", "sac": "sac", "no_computable": "no_computable",
            "aporte_real": "ap_jub", "sindical": "sindical", "otros_desc": "otros_desc",
            "alq_beneficio": "alq_beneficio", "ded_grales_otras": "ded_grales_otras",
            "cajas_compl": "cajas_compl", "conyuge": "conyuge", "hijos": "hijos",
            "hijo_incap": "hijo_incap", "ret_informada": "ret_informada"}
    for i in range(n):
        wd.cell(5 + i, 2, i + 1).border = thin
        for j, campo in enumerate(CAMPOS):
            v = datos[mapa[campo]][i]
            cc = wd.cell(5 + i, 3 + j, v)
            cc.border = thin
            if isinstance(v, (int, float)):
                cc.number_format = money
                cc.alignment = R
    wd.column_dimensions["B"].width = 8
    for j in range(len(CAMPOS)):
        wd.column_dimensions[get_column_letter(3 + j)].width = 15

    # ---- Parametros ----
    wp = wb.create_sheet("Parametros")
    wp.sheet_view.showGridLines = False
    wp["B2"] = "PARAMETROS DEL PERIODO FISCAL (params.json)"
    wp["B2"].font = H1
    wp.cell(4, 2, "Deducciones personales art. 30 ACUMULADAS por mes de pago").font = BOLD
    labs = ["Ganancia no imponible", "Conyuge", "Hijo", "Hijo incapacitado", "Deduccion especial ap.2"]
    concs = ["MNI", "CONYUGE", "HIJO", "HIJO_INCAP", "DED_ESP"]
    wp.cell(5, 2, "Mes pago").font = H2
    wp.cell(5, 2).fill = FILL
    for j, l in enumerate(labs):
        cc = wp.cell(5, 3 + j, l)
        cc.font = H2
        cc.fill = FILL
        cc.alignment = WRAP
    for m in range(1, 13):
        wp.cell(5 + m, 2, MESES[m - 1].capitalize()).border = thin
        for j, cn in enumerate(concs):
            cc = wp.cell(5 + m, 3 + j, round(dp_acumulada(cn, m), 2))
            cc.number_format = money
            cc.alignment = R
            cc.border = thin
    r0 = 20
    wp.cell(r0, 2, "Topes mensuales base imponible aportes SS (ANSES)").font = BOLD
    wp.cell(r0 + 1, 2, "Mes").font = H2
    wp.cell(r0 + 1, 3, "Tope").font = H2
    for m in range(12):
        wp.cell(r0 + 2 + m, 2, MESES[m].capitalize())
        c = wp.cell(r0 + 2 + m, 3, TOPE_SS[m])
        c.number_format = money
        c.alignment = R
    r0 = r0 + 2 + 12 + 2
    wp.cell(r0, 2, "Escala art. 94 - MENSUAL ACUMULADA por mes de pago").font = BOLD
    for j, h in enumerate(["Mes", "Desde", "Hasta", "Fijo", "Alicuota", "Excedente"]):
        wp.cell(r0 + 1, 2 + j, h).font = H2
    rr = r0 + 2
    for m in range(12):
        for (d, h, f, a, e) in SCALES[ABREV[m]]:
            wp.cell(rr, 2, MESES[m].capitalize())
            wp.cell(rr, 3, round(d, 2)).number_format = money
            wp.cell(rr, 4, "en adelante" if h is None else round(h, 2))
            wp.cell(rr, 5, round(f, 2)).number_format = money
            wp.cell(rr, 6, a).number_format = "0%"
            wp.cell(rr, 7, round(e, 2)).number_format = money
            rr += 1
    wp.column_dimensions["B"].width = 16
    for cl in "CDEFG":
        wp.column_dimensions[cl].width = 18

    wb.save(xlsx_path)
    print("Planilla generada:", xlsx_path)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    argv = sys.argv[3:]
    armar(
        sys.argv[1], sys.argv[2],
        mes_siguiente=("--mes-siguiente" in argv),
        prorrateo_sac=("--con-prorrateo-sac" in argv),
        aportes_tope=("--aportes-tope" in argv),
        conyuge=("--conyuge" in argv),
        hijos=("--hijos" in argv),
        hijo_incap=("--hijo-incap" in argv),
    )
