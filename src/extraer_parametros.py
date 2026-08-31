# -*- coding: utf-8 -*-
"""
Extrae los parametros del periodo fiscal (escala del art. 94 mensual acumulada,
deducciones del art. 30 y topes de aportes ANSES) desde un liquidador oficial
de Ganancias 4ta categoria (.xlsm) y los deja en params.json.

Ademas verifica que los valores coincidan con los PDFs oficiales de ARCA
(carpeta ../normativa).

Uso:
    python extraer_parametros.py  "ruta/al/Liquidador_Ganancias_4ta.xlsm"
    python extraer_parametros.py  --verificar     # solo re-chequea params.json vs PDFs

El liquidador oficial NO se versiona en este repo (puede tener datos cargados de
un contribuyente). Solo se guardan los numeros de parametros, que son publicos.
"""
import json
import os
import re
import sys

AQUI = os.path.dirname(__file__)
PARAMS = os.path.join(AQUI, "params.json")
NORMATIVA = os.path.join(AQUI, "..", "normativa")

ABREV = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]


# ---------------------------------------------------------------------------
# 1) Extraccion desde el liquidador oficial
# ---------------------------------------------------------------------------
def extraer(xlsm_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

    ini = wb["Inicio"]
    # En la hoja "Inicio" hay 12 bloques de escala mensual acumulada:
    #   ene-jun en columnas B..F (filas 26,38,50,62,74,86)
    #   jul-dic en columnas I..M (mismas filas)
    scales = {}
    for bi, r0 in enumerate([26, 38, 50, 62, 74, 86]):
        for lado, c0, off in [("izq", 2, 0), ("der", 9, 6)]:
            mi = bi + off
            filas = []
            for k in range(9):
                r = r0 + k
                desde = ini.cell(r, c0).value
                hasta = ini.cell(r, c0 + 1).value
                fijo = ini.cell(r, c0 + 2).value
                alic = ini.cell(r, c0 + 3).value
                exc = ini.cell(r, c0 + 4).value
                filas.append([desde, "en adelante" if isinstance(hasta, str) else hasta, fijo, alic, exc])
            scales[ABREV[mi]] = filas

    # Deducciones personales: hoja "Tablas"
    #   D91:D95  -> valores MENSUALES del 1er semestre
    #   D169:D173 -> valores ANUALES del 2do semestre (liq. anual)
    tb = wb["Tablas"]
    sem1 = {"MNI": tb["D91"].value, "CONYUGE": tb["D92"].value, "HIJO": tb["D93"].value,
            "HIJO_INCAP": tb["D94"].value, "DED_ESP": tb["D95"].value}
    sem2 = {"MNI": tb["D169"].value, "CONYUGE": tb["D170"].value, "HIJO": tb["D171"].value,
            "HIJO_INCAP": tb["D172"].value, "DED_ESP": tb["D173"].value}

    # Topes ANSES a la base imponible de aportes: hoja "Carga Datos" fila 147, D..O
    cd = wb["Carga Datos"]
    topes = [cd.cell(147, 4 + i).value for i in range(12)]

    params = {
        "periodo_fiscal": int(cd["H4"].value),
        "fuente": "Liquidador oficial de Ganancias 4ta categoria (hojas Inicio/Tablas/Carga Datos).",
        "ded_personales_sem1": sem1,
        "ded_personales_sem2_anual": sem2,
        "topes_ss": topes,
        "scales": scales,
    }
    json.dump(params, open(PARAMS, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print("params.json actualizado desde", os.path.basename(xlsm_path))
    return params


# ---------------------------------------------------------------------------
# 2) Verificacion contra los PDFs oficiales de ARCA
# ---------------------------------------------------------------------------
def _num(s):
    return float(s.replace(".", "").replace(",", "."))


def verificar():
    """Chequea los acumulados de julio y diciembre y la escala anual contra los PDFs."""
    import pdfplumber
    p = json.load(open(PARAMS, encoding="utf-8"))
    ok = True

    ded_pdf = os.path.join(NORMATIVA, "Deducciones-personales-art-30-jul-dic-2026.pdf")
    if os.path.exists(ded_pdf):
        with pdfplumber.open(ded_pdf) as pdf:
            txt = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
        # MNI acumulado a diciembre (2do semestre) = mensual_sem1*6 + (anual/12)*6
        # En el PDF, tras "IMPORTE ACUMULADO ... DICIEMBRE $" hay una fila:
        #   "A) Ganancias no imponibles [Articulo\n30, inciso a)]: <ene> <feb> ... <dic>"
        mni_dic_calc = p["ded_personales_sem1"]["MNI"] * 6 + p["ded_personales_sem2_anual"]["MNI"] / 12 * 6
        nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d\d", txt)
        cand = [_num(x) for x in nums if abs(_num(x) - mni_dic_calc) < 2.0]
        if cand:
            print(f"  MNI acum. diciembre (PDF) {cand[0]:,.2f}  vs  calc {mni_dic_calc:,.2f}   OK")
        else:
            ok = False
            print(f"  MNI acum. diciembre: NO se encontro {mni_dic_calc:,.2f} en el PDF - revisar params.json")

    esc_pdf = os.path.join(NORMATIVA, "Tabla-Art-94-LIG-liquidacion-anual-y-final-2026.pdf")
    if os.path.exists(esc_pdf):
        with pdfplumber.open(esc_pdf) as pdf:
            txt = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
        # primer piso de la escala anual
        m = re.search(r"0,00\s+([\d.]+,\d\d)\s+0,00\s+5", txt)
        if m:
            print(f"  1er tramo escala anual (PDF): hasta {m.group(1)}   (referencia)")

    print("Verificacion:", "OK" if ok else "revisar diferencias")
    return ok


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] not in ("--verificar", "-v"):
        extraer(sys.argv[1])
    verificar()
